
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib
import numpy as np
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    precision_recall_fscore_support,
    classification_report,
    confusion_matrix
)
import warnings
from openpyxl import load_workbook
import argparse

warnings.filterwarnings('ignore')

# 中文字体设置
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'FangSong', 'KaiTi', 'Arial Unicode MS']
matplotlib.rcParams['axes.unicode_minus'] = False
matplotlib.rcParams['font.size'] = 10


def process_classification_report(y_true, y_pred, all_labels):
    """处理分类报告，删除accuracy行并处理零样本情况"""

    # 获取分类报告字典
    report_dict = classification_report(y_true, y_pred, labels=all_labels, zero_division=0, output_dict=True)

    # 删除 accuracy 行
    if 'accuracy' in report_dict:
        del report_dict['accuracy']

    # 转换为DataFrame
    report_df = pd.DataFrame(report_dict).transpose()

    # 重命名support列为total_samples
    report_df = report_df.rename(columns={'support': 'total_samples'})

    # 处理零样本情况：support为0时，precision、recall、f1用"-"代替
    zero_support_mask = (report_df['total_samples'] == 0)

    for col in ['precision', 'recall', 'f1-score']:
        if col in report_df.columns:
            report_df.loc[zero_support_mask, col] = '-'

    # 格式化其他数值列为4位小数（排除已经是"-"的列和total_samples）
    for col in report_df.columns:
        if col != 'total_samples':
            # 只对数值类型的列进行格式化
            numeric_mask = pd.to_numeric(report_df[col], errors='coerce').notna()
            if numeric_mask.any():
                report_df.loc[numeric_mask, col] = pd.to_numeric(report_df.loc[numeric_mask, col]).round(4)

    return report_df


def print_classification_report_custom(y_true, y_pred, all_labels, label_width=15):
    """自定义打印分类报告，不包含accuracy行"""

    report_dict = classification_report(y_true, y_pred, labels=all_labels, zero_division=0, output_dict=True)

    # 删除accuracy行
    filtered_report_dict = {k: v for k, v in report_dict.items() if k != 'accuracy'}

    # 打印表头
    print(f"{'':>{label_width}} {'precision':>9} {'recall':>9} {'f1-score':>9} {'support':>9}")
    print()

    # 打印各类别指标
    for label in all_labels:
        if label in filtered_report_dict:
            metrics = filtered_report_dict[label]
            support = int(metrics['support'])

            # 如果support为0，显示"-"
            if support == 0:
                print(f"{label:>{label_width}} {'-':>9} {'-':>9} {'-':>9} {support:>9}")
            else:
                print(
                    f"{label:>{label_width}} {metrics['precision']:>9.2f} {metrics['recall']:>9.2f} {metrics['f1-score']:>9.2f} {support:>9}")

    print()
    # 显示平均值
    for avg_type in ['macro avg', 'weighted avg']:
        if avg_type in filtered_report_dict:
            metrics = filtered_report_dict[avg_type]
            support = int(metrics['support'])
            print(
                f"{avg_type:>{label_width}} {metrics['precision']:>9.2f} {metrics['recall']:>9.2f} {metrics['f1-score']:>9.2f} {support:>9}")


# 领域分类评估
def evaluate_domain_classification(df, results_dict):
    print("\n" + "=" * 80)
    print("任务1：领域分类评估")
    print("=" * 80)

    y_true = df['raw_domain']
    y_pred = df['domain']

    # 基础统计
    print(f"\n数据总量: {len(df)} 条")
    print(f"类别数量: {y_true.nunique()} 个 ({', '.join(sorted(y_true.unique()))})")

    # 整体指标
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall = recall_score(y_true, y_pred, average='macro', zero_division=0)
    f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)


    print(f"\n【整体指标】")
    print(f"准确率 (Accuracy): {accuracy:.4f} ({accuracy * 100:.2f}%)")
    print(f"精确率 (Precision): {precision:.4f} ({precision * 100:.2f}%)")
    print(f"召回率 (Recall): {recall:.4f} ({recall * 100:.2f}%)")
    print(f"f1 (F1): {f1:.4f} ({f1 * 100:.2f}%)")


    # 各类别详细指标
    print(f"\n【各类别详细指标】")
    # 获取所有出现的标签（真实标签和预测标签的并集）
    all_labels = sorted(set(y_true.unique()) | set(y_pred.unique()))

    # 打印自定义的分类报告（不包含accuracy行）
    print_classification_report_custom(y_true, y_pred, all_labels, label_width=15)

    # 处理DataFrame保存
    report_df = process_classification_report(y_true, y_pred, all_labels)

    # 缓存领域分类-详细指标, 保存进sheet
    results_dict['领域分类-详细指标'] = report_df

    # 错误分析
    print(f"\n【错误样本统计】(按数量降序)")
    error_df = df[y_true != y_pred]
    if len(error_df) > 0:
        error_stats = error_df.groupby(['raw_domain', 'domain']).size().reset_index(name='错误数量')
        error_stats = error_stats.sort_values('错误数量', ascending=False)

        # 保存错误统计到Excel
        results_dict['领域分类-错误统计'] = error_stats

        print(f"总错误数: {len(error_df)} 条 (错误率: {len(error_df) / len(df) * 100:.2f}%)\n")
        for idx, row in error_stats.iterrows():
            print(f"{row['raw_domain']} → {row['domain']}: {row['错误数量']}个")
    else:
        print("完美分类！无错误样本。")
        # 即使没有错误，也保存一个空的DataFrame
        results_dict['领域分类-错误统计'] = pd.DataFrame(columns=['raw_domain', 'domain', '错误数量'])
    # 可视化
    visualize_domain_classification(y_true, y_pred)


def visualize_domain_classification(y_true, y_pred):
    """领域分类可视化"""

    # 图1：混淆矩阵热力图
    fig, ax = plt.subplots(figsize=(10, 8))

    # 🔧 修改：使用真实+预测的并集，与classification_report保持一致
    labels = sorted(set(y_true.unique()) | set(y_pred.unique()))
    cm = confusion_matrix(y_true, y_pred, labels=labels)

    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                xticklabels=labels, yticklabels=labels,
                cbar_kws={'label': '样本数量'}, ax=ax)

    ax.set_title('领域分类混淆矩阵', fontsize=16, fontweight='bold', pad=20)
    ax.set_xlabel('预测领域', fontsize=13, fontweight='bold')
    ax.set_ylabel('真实领域', fontsize=13, fontweight='bold')

    ax.set_xticklabels(ax.get_xticklabels())
    ax.set_yticklabels(ax.get_yticklabels())

    plt.tight_layout()
    plt.show()

    # 图2：各类别指标对比柱状图
    # 🔧 修改：使用相同的标签集合
    precision, recall, f1, support = precision_recall_fscore_support(
        y_true, y_pred, labels=labels, zero_division=0
    )

    fig, ax = plt.subplots(figsize=(12, 7))

    x = np.arange(len(labels))
    width = 0.25

    bars1 = ax.bar(x - width, precision, width, label='Precision', color='#E74C3C', alpha=0.8)
    bars2 = ax.bar(x, recall, width, label='Recall', color='#3498DB', alpha=0.8)
    bars3 = ax.bar(x + width, f1, width, label='F1-Score', color='#2ECC71', alpha=0.8)

    ax.set_xlabel('领域', fontsize=13, fontweight='bold')
    ax.set_ylabel('分数', fontsize=13, fontweight='bold')
    ax.set_title('各领域分类指标对比', fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(fontsize=11)
    ax.set_ylim(0, 1.1)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 添加数值标签
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{height:.2f}',
                    ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    plt.show()

    # 图3：预测分布 vs 真实分布对比
    fig, ax = plt.subplots(figsize=(12, 7))

    true_counts = y_true.value_counts().reindex(labels, fill_value=0)
    pred_counts = y_pred.value_counts().reindex(labels, fill_value=0)

    x = np.arange(len(labels))
    width = 0.35

    bars1 = ax.bar(x - width / 2, true_counts.values, width, label='真实分布',
                   color='#66C2A5', alpha=0.8, edgecolor='black', linewidth=1)
    bars2 = ax.bar(x + width / 2, pred_counts.values, width, label='预测分布',
                   color='#FC8D62', alpha=0.8, edgecolor='black', linewidth=1)

    ax.set_xlabel('领域', fontsize=13, fontweight='bold')
    ax.set_ylabel('样本数量', fontsize=13, fontweight='bold')
    ax.set_title('预测分布 vs 真实分布对比', fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(fontsize=11)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 添加数值标签
    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{int(height)}',
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    plt.tight_layout()
    plt.show()


# 领域-意图联合分类评估
def evaluate_domain_intent_classification(df, results_dict):
    print("\n" + "=" * 80)
    print("任务2：领域-意图联合分类评估")
    print("=" * 80)

    # 创建组合标签
    y_true = df['raw_domain'] + '-' + df['raw_intent']
    y_pred = df['domain'] + '-' + df['intent']

    # 基础统计
    print(f"\n数据总量: {len(df)} 条")
    print(f"组合类别数量: {y_true.nunique()} 个")

    # 整体指标
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall = recall_score(y_true, y_pred, average='macro', zero_division=0)
    f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)

    print(f"\n【整体指标】")
    print(f"准确率 (Accuracy): {accuracy:.4f} ({accuracy * 100:.2f}%)")
    print(f"精确率 (Precision): {precision:.4f} ({precision * 100:.2f}%)")
    print(f"召回率 (Recall): {recall:.4f} ({recall * 100:.2f}%)")
    print(f"f1 (F1): {f1:.4f} ({f1 * 100:.2f}%)")

    # 各组合类别详细指标
    print(f"\n【各组合类别详细指标】")

    # 获取所有出现的标签（真实标签和预测标签的并集）
    all_labels = sorted(set(y_true.unique()) | set(y_pred.unique()))

    # 打印自定义的分类报告（不包含accuracy行）
    print_classification_report_custom(y_true, y_pred, all_labels, label_width=25)

    # 处理DataFrame保存
    report_df = process_classification_report(y_true, y_pred, all_labels)

    results_dict['领域意图联合-详细指标'] = report_df

    # 分层分析
    print(f"\n【分层分析】")
    # 领域层面
    domain_correct = (df['raw_domain'] == df['domain']).sum()
    domain_accuracy = domain_correct / len(df)
    print(f"领域正确率（不考虑意图）: {domain_accuracy:.4f} ({domain_accuracy * 100:.2f}%)")

    # 各领域内的意图正确率
    for domain in sorted(df['raw_domain'].unique()):
        domain_df = df[df['raw_domain'] == domain]
        if len(domain_df) > 0:
            # 领域预测也对的情况下，意图的正确率
            domain_correct_df = domain_df[domain_df['raw_domain'] == domain_df['domain']]
            if len(domain_correct_df) > 0:
                intent_correct = (domain_correct_df['raw_intent'] == domain_correct_df['intent']).sum()
                intent_accuracy = intent_correct / len(domain_correct_df)
                print(
                    f"  {domain}领域（领域预测正确的前提下）: 意图正确率 {intent_accuracy:.4f} ({intent_accuracy * 100:.2f}%)")

    print(f"\n完全匹配（领域+意图都正确）: {accuracy:.4f} ({accuracy * 100:.2f}%)")

    # 错误分析
    print(f"\n【错误样本统计】(按数量降序)")
    error_df = df[y_true != y_pred]
    if len(error_df) > 0:
        error_df_with_labels = error_df.copy()
        error_df_with_labels['true_label'] = y_true[y_true != y_pred]
        error_df_with_labels['pred_label'] = y_pred[y_true != y_pred]

        error_stats = error_df_with_labels.groupby(['true_label', 'pred_label']).size().reset_index(name='错误数量')
        error_stats = error_stats.sort_values('错误数量', ascending=False)

        # 保存错误统计到Excel
        results_dict['领域意图联合-错误统计'] = error_stats

        print(f"总错误数: {len(error_df)} 条 (错误率: {len(error_df) / len(df) * 100:.2f}%)\n")
        print(f"前10个最常见的错误类型:")
        for idx, row in error_stats.head(10).iterrows():
            print(f"{row['true_label']} → {row['pred_label']}: {row['错误数量']}个")
    else:
        print("完美分类！无错误样本。")
        # 即使没有错误，也保存一个空的DataFrame
        results_dict['领域意图联合-错误统计'] = pd.DataFrame(columns=['true_label', 'pred_label', '错误数量'])

    # 可视化
    visualize_domain_intent_classification(y_true, y_pred, df)


def visualize_domain_intent_classification(y_true, y_pred, df):
    """领域-意图联合分类可视化"""

    # 图1：混淆矩阵热力图
    # 🔧 修改：使用真实+预测的并集，与classification_report保持一致
    labels = sorted(set(y_true.unique()) | set(y_pred.unique()))

    if len(labels) <= 15:
        fig, ax = plt.subplots(figsize=(14, 12))

        cm = confusion_matrix(y_true, y_pred, labels=labels)

        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=labels, yticklabels=labels,
                    cbar_kws={'label': '样本数量'}, ax=ax)

        ax.set_title('领域-意图联合分类混淆矩阵', fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel('预测类别', fontsize=13, fontweight='bold')
        ax.set_ylabel('真实类别', fontsize=13, fontweight='bold')

        ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right', fontsize=9)
        ax.set_yticklabels(ax.get_yticklabels(), rotation=0, fontsize=9)

        plt.tight_layout()
        plt.show()
    else:
        print(f"\n[提示] 类别数量较多({len(labels)}个)，混淆矩阵过大，跳过完整混淆矩阵可视化")

    # 图2：各组合类别准确率排行
    # 🔧 使用相同的标签集合
    # 计算每个类别的准确率
    accuracy_list = []
    support_list = []

    for label in labels:
        # 找到该类别的所有样本
        mask = (y_true == label)
        total = mask.sum()

        if total > 0:
            # 计算该类别预测正确的数量
            correct = ((y_true == label) & (y_pred == label)).sum()
            accuracy = correct / total
        else:
            accuracy = 0

        accuracy_list.append(accuracy)
        support_list.append(total)

    # 创建DataFrame便于排序
    metrics_df = pd.DataFrame({
        '类别': labels,
        '准确率': accuracy_list,
        'Support': support_list
    }).sort_values('准确率', ascending=True)  # 升序，低分在前

    # 只显示前20个
    display_df = metrics_df.tail(20) if len(metrics_df) > 20 else metrics_df

    fig, ax = plt.subplots(figsize=(10, max(8, len(display_df) * 0.4)))

    colors = ['#E74C3C' if score < 0.5 else '#F39C12' if score < 0.7 else '#2ECC71'
              for score in display_df['准确率']]

    bars = ax.barh(range(len(display_df)), display_df['准确率'], color=colors, alpha=0.8)

    ax.set_yticks(range(len(display_df)))
    ax.set_yticklabels(display_df['类别'], fontsize=9)
    ax.set_xlabel('准确率', fontsize=12, fontweight='bold')
    ax.set_title(f'领域-意图准确率排行',
                 fontsize=14, fontweight='bold', pad=15)
    ax.set_xlim(0, 1.1)
    ax.grid(axis='x', alpha=0.3, linestyle='--')

    # 添加数值标签
    for i, (bar, score, sup) in enumerate(zip(bars, display_df['准确率'], display_df['Support'])):
        ax.text(score + 0.02, i, f'{score:.3f} (n={sup})',
                va='center', fontsize=8)

    plt.tight_layout()
    plt.show()

    # 图3：按领域分组的意图准确率对比
    fig, ax = plt.subplots(figsize=(12, 7))

    domains = sorted(df['raw_domain'].unique())
    intent_accuracies = []

    for domain in domains:
        domain_df = df[df['raw_domain'] == domain]
        domain_correct_df = domain_df[domain_df['raw_domain'] == domain_df['domain']]
        if len(domain_correct_df) > 0:
            intent_correct = (domain_correct_df['raw_intent'] == domain_correct_df['intent']).sum()
            intent_accuracy = intent_correct / len(domain_correct_df)
        else:
            intent_accuracy = 0
        intent_accuracies.append(intent_accuracy)

    colors_domain = ['#66C2A5', '#FC8D62', '#8DA0CB', '#E78AC3']
    bars = ax.bar(range(len(domains)), intent_accuracies,
                  color=colors_domain[:len(domains)], alpha=0.8,
                  edgecolor='black', linewidth=1.2)

    ax.set_xticks(range(len(domains)))
    ax.set_xticklabels(domains, fontsize=12)
    ax.set_ylabel('意图正确率', fontsize=13, fontweight='bold')
    ax.set_title('各领域内的意图分类准确率\n(仅统计领域预测正确的样本)',
                 fontsize=14, fontweight='bold', pad=20)
    ax.set_ylim(0, 1.1)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 添加数值标签
    for i, (bar, acc) in enumerate(zip(bars, intent_accuracies)):
        domain_df = df[df['raw_domain'] == domains[i]]
        domain_correct_df = domain_df[domain_df['raw_domain'] == domain_df['domain']]
        ax.text(i, acc + 0.02, f'{acc:.2%}\n(n={len(domain_correct_df)})',
                ha='center', va='bottom', fontsize=10, fontweight='bold')

    plt.tight_layout()
    plt.show()



# 和解状态分类评估
def evaluate_reconciliation_classification(df, results_dict):
    print("\n" + "=" * 80)
    print("任务3：和解状态分类评估")
    print("=" * 80)

    y_true = df['real_reconciliation']
    y_pred = df['reconciliation']

    # 基础统计
    print(f"\n数据总量: {len(df)} 条")
    print(f"类别数量: {y_true.nunique()} 个 ({', '.join(sorted(y_true.unique()))})")

    # 整体指标
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall = recall_score(y_true, y_pred, average='macro', zero_division=0)
    f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)


    print(f"\n【整体指标】")
    print(f"准确率 (Accuracy): {accuracy:.4f} ({accuracy * 100:.2f}%)")
    print(f"精确率 (Precision): {precision:.4f} ({precision * 100:.2f}%)")
    print(f"召回率 (Recall): {recall:.4f} ({recall * 100:.2f}%)")
    print(f"f1 (F1): {f1:.4f} ({f1 * 100:.2f}%)")


    # 各类别详细指标
    print(f"\n【各类别详细指标】")
    # 获取所有出现的标签（真实标签和预测标签的并集）
    all_labels = sorted(set(y_true.unique()) | set(y_pred.unique()))

    # 打印自定义的分类报告（不包含accuracy行）
    print_classification_report_custom(y_true, y_pred, all_labels, label_width=15)

    # 处理DataFrame保存
    report_df = process_classification_report(y_true, y_pred, all_labels)

    # 缓存领域分类-详细指标, 保存进sheet
    results_dict['和解状态分类-详细指标'] = report_df

    # 错误分析
    print(f"\n【错误样本统计】(按数量降序)")
    error_df = df[y_true != y_pred]
    if len(error_df) > 0:
        error_stats = error_df.groupby(['real_reconciliation', 'reconciliation']).size().reset_index(name='错误数量')
        error_stats = error_stats.sort_values('错误数量', ascending=False)

        # 保存错误统计到Excel
        results_dict['和解状态分类-错误统计'] = error_stats

        print(f"总错误数: {len(error_df)} 条 (错误率: {len(error_df) / len(df) * 100:.2f}%)\n")
        for idx, row in error_stats.iterrows():
            print(f"{row['real_reconciliation']} → {row['reconciliation']}: {row['错误数量']}个")
    else:
        print("完美分类！无错误样本。")
        # 即使没有错误，也保存一个空的DataFrame
        results_dict['和解状态分类-错误统计'] = pd.DataFrame(columns=['real_reconciliation', 'reconciliation', '错误数量'])
    # 可视化
    visualize_reconciliation_classification(y_true, y_pred)

def visualize_reconciliation_classification(y_true, y_pred):
    """和解状态分类可视化"""

    # 图1：混淆矩阵热力图
    fig, ax = plt.subplots(figsize=(10, 8))

    # 🔧 修改：使用真实+预测的并集，与classification_report保持一致
    labels = sorted(set(y_true.unique()) | set(y_pred.unique()))
    cm = confusion_matrix(y_true, y_pred, labels=labels)

    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                xticklabels=labels, yticklabels=labels,
                cbar_kws={'label': '样本数量'}, ax=ax)

    ax.set_title('和解状态分类混淆矩阵', fontsize=16, fontweight='bold', pad=20)
    ax.set_xlabel('预测和解状态', fontsize=13, fontweight='bold')
    ax.set_ylabel('真实和解状态', fontsize=13, fontweight='bold')

    ax.set_xticklabels(ax.get_xticklabels())
    ax.set_yticklabels(ax.get_yticklabels())

    plt.tight_layout()
    plt.show()

    # 图2：各类别指标对比柱状图
    # 🔧 修改：使用相同的标签集合
    precision, recall, f1, support = precision_recall_fscore_support(
        y_true, y_pred, labels=labels, zero_division=0
    )

    fig, ax = plt.subplots(figsize=(12, 7))

    x = np.arange(len(labels))
    width = 0.25

    bars1 = ax.bar(x - width, precision, width, label='Precision', color='#E74C3C', alpha=0.8)
    bars2 = ax.bar(x, recall, width, label='Recall', color='#3498DB', alpha=0.8)
    bars3 = ax.bar(x + width, f1, width, label='F1-Score', color='#2ECC71', alpha=0.8)

    ax.set_xlabel('和解状态', fontsize=13, fontweight='bold')
    ax.set_ylabel('分数', fontsize=13, fontweight='bold')
    ax.set_title('和解状态分类指标对比', fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(fontsize=11)
    ax.set_ylim(0, 1.1)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 添加数值标签
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{height:.2f}',
                    ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    plt.show()

    # 图3：预测分布 vs 真实分布对比
    fig, ax = plt.subplots(figsize=(12, 7))

    true_counts = y_true.value_counts().reindex(labels, fill_value=0)
    pred_counts = y_pred.value_counts().reindex(labels, fill_value=0)

    x = np.arange(len(labels))
    width = 0.35

    bars1 = ax.bar(x - width / 2, true_counts.values, width, label='真实分布',
                   color='#66C2A5', alpha=0.8, edgecolor='black', linewidth=1)
    bars2 = ax.bar(x + width / 2, pred_counts.values, width, label='预测分布',
                   color='#FC8D62', alpha=0.8, edgecolor='black', linewidth=1)

    ax.set_xlabel('和解状态', fontsize=13, fontweight='bold')
    ax.set_ylabel('样本数量', fontsize=13, fontweight='bold')
    ax.set_title('预测分布 vs 真实分布对比', fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(fontsize=11)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # 添加数值标签
    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{int(height)}',
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    plt.tight_layout()
    plt.show()


def save_results_to_excel(file_path, results_dict, original_df):
    """将结果保存到Excel文件的新sheet中"""
    print("\n" + "=" * 80)
    print("开始保存结果到Excel...")
    print("=" * 80)

    try:
        # 先读取原始数据，保留所有现有sheet
        with pd.ExcelFile(file_path) as xls:
            existing_sheets = {}
            for sheet_name in xls.sheet_names:
                # 如果是第一个sheet（原始数据），用更新后的df替换
                if sheet_name == xls.sheet_names[0]:
                    existing_sheets[sheet_name] = original_df
                else:
                    existing_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)

        # 使用ExcelWriter写入（覆盖模式）
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 先写入原有的sheets（包括更新后的原始数据）
            for sheet_name, sheet_df in existing_sheets.items():
                sheet_df.to_excel(writer, sheet_name='原始数据', index=False)

            # 写入新的结果sheets
            for sheet_name, result_df in results_dict.items():
                # 错误统计表不需要索引
                if '错误统计' in sheet_name:
                    result_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    result_df.to_excel(writer, sheet_name=sheet_name, index=True)
                print(f"✓ 已保存表格: {sheet_name}")

        # 使用openpyxl格式化数值列的小数位数
        wb = load_workbook(file_path)

        for sheet_name in results_dict.keys():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 错误统计表不需要格式化
                if '错误统计' in sheet_name:
                    continue

                # 从第2行开始（第1行是表头），第2列开始（第1列是索引）
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        # 如果是数值类型，设置格式
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            # 检查列名是否为total_samples
                            col_name = ws.cell(1, cell.column).value
                            if col_name == 'total_samples':
                                cell.number_format = '0'
                            else:
                                cell.number_format = '0.0000'
                        # 如果是"-"字符串，保持不变
                        elif cell.value == '-':
                            pass  # 保持为文本格式

        wb.save(file_path)
        wb.close()

        print("\n" + "=" * 80)
        print(f"✅ 所有结果已成功保存到: {file_path}")
        print(f"   - 原始数据已更新")
        print(f"   - 新增 {len(results_dict)} 个评估指标表")
        print("=" * 80)

    except Exception as e:
        print(f"\n❌ 保存过程中出现错误: {str(e)}")
        print("提示: 请确保Excel文件没有被其他程序打开")

def main():
    results_to_save = {}

    parser = argparse.ArgumentParser(description='模型打标结果评估脚本')
    parser.add_argument('file_path', help='需要评估的文件路径(评估结果同步保存在源文件中)')
    parser.add_argument('--claim_point', action='store_true', help='诉点评估')
    parser.add_argument('--reconciliation_state', action='store_true', help='和解状态评估')

    args = parser.parse_args()
    if not args.claim_point and not args.reconciliation_state:
        print("\n⚠️  未指定任何评估任务!")
        print("使用 --claim_point 启用诉点评估")
        print("使用 --reconciliation_state 启用和解状态评估")
        print("可以同时使用两个参数")
    # 读取数据
    df = pd.read_excel(args.file_path)
    df = df[df['flag'] == 'qualified']
    print(df.head())


    print("=" * 80)
    print("数据加载完成")
    print("=" * 80)
    print(f"总数据量: {len(df)} 条")
    print(f"列名: {', '.join(df.columns.tolist())}")


    if args.claim_point:

        # ========== 添加两个新列 ==========
        df['domain_correct'] = (df['raw_domain'] == df['domain']).astype(int)
        df['domain_intent_correct'] = ((df['raw_domain'] + '-' + df['raw_intent']) ==
                                       (df['domain'] + '-' + df['intent'])).astype(int)
        print(f"\n前5行数据预览:")
        print(df[['raw_domain', 'domain', 'raw_intent', 'intent', 'domain_correct', 'domain_intent_correct']].head())
        # 领域分类评估
        evaluate_domain_classification(df, results_to_save)

        # 领域-意图联合分类评估
        evaluate_domain_intent_classification(df, results_to_save)



    if args.reconciliation_state:
        df['reconciliation_correct'] = (df['real_reconciliation'] == df['reconciliation']).astype(int)
        print(f"\n前5行数据预览:")
        print(df[['real_reconciliation', 'reconciliation',  'reconciliation_correct']].head())
        # 和解方案评估
        evaluate_reconciliation_classification(df, results_to_save)

    # 保存结果到Excel（包括更新后的原始数据和评估指标表）
    if results_to_save:
        save_results_to_excel(args.file_path, results_to_save, df)


# 主程序执行
if __name__ == "__main__":
    main()
